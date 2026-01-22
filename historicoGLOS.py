# app.py
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode

import pandas as pd
import plotly.express as px
import requests
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------
# STREAMLIT CONFIG
# ----------------------------------------------------
st.set_page_config(
    page_title="Bonsaif m=27 — Tipificación / Calificación / Colgó (GLOS)",
    page_icon="📞",
    layout="wide",
)

# ----------------------------------------------------
# HELPERS
# ----------------------------------------------------
def _strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(ch))


def _norm_colname(s: str) -> str:
    s = str(s).strip()
    s = _strip_accents(s)
    s = s.replace(" ", "_")
    return s.lower()


def first_existing_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    if df is None or df.empty:
        return None
    cols = list(df.columns)
    norm_map = {_norm_colname(c): c for c in cols}
    for cand in candidates:
        key = _norm_colname(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def _pick_records_from_payload(payload):
    if payload is None:
        return []
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        preferred_keys = ["data", "datos", "rows", "result", "resultado", "reporte", "items"]
        for k in preferred_keys:
            if k in payload and isinstance(payload[k], list):
                return payload[k]
        for _, v in payload.items():
            if isinstance(v, list) and (len(v) == 0 or isinstance(v[0], dict)):
                return v
    return []


def _payload_message(payload):
    if payload is None:
        return None
    if isinstance(payload, str):
        return payload.strip() if payload.strip() else None
    if isinstance(payload, dict):
        for k in ["mensaje", "message", "msg", "error", "detalle"]:
            if k in payload and isinstance(payload[k], str) and payload[k].strip():
                return payload[k].strip()
    return None


def _clean_text_to_na(series: pd.Series) -> pd.Series:
    """
    Convert typical 'empty' strings to NA so we never display 'nan',
    then you can fillna("SIN_SUPERVISOR") and it WILL show.
    """
    s = series.astype("string")
    s = s.str.strip()
    s = s.replace(
        {
            "": pd.NA,
            "nan": pd.NA,
            "NaN": pd.NA,
            "None": pd.NA,
            "NULL": pd.NA,
            "null": pd.NA,
            "N/A": pd.NA,
            "n/a": pd.NA,
        }
    )
    return s


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    dt_col = first_existing_col(df, ["Fecha_CC", "Fecha", "fecha", "fecha_cc"])
    if dt_col:
        df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce")
        df["Fecha_CC"] = df[dt_col]
        df["Dia"] = df["Fecha_CC"].dt.date

    for c in ["Duracion_CC", "Duración_Min_CC", "Duracion_Min_CC", "acw", "Extension_CC", "Codigo_Sip_CC"]:
        col = first_existing_col(df, [c])
        if col:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    dur_sec = first_existing_col(df, ["Duracion_CC"])
    dur_min = first_existing_col(df, ["Duracion_Min_CC", "Duración_Min_CC"])
    if dur_sec and not dur_min:
        df["Duracion_Min_CC"] = (pd.to_numeric(df[dur_sec], errors="coerce") / 60.0).round(2)

    return df


def to_excel_bytes(df: pd.DataFrame, sheet_name="m27", max_autofit_rows: int = 200) -> bytes:
    """
    Safer Excel export to avoid Streamlit Cloud crashes:
    - Auto-fit widths using ONLY header + first N rows (NOT the whole dataset).
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb[sheet_name]

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ✅ Auto-fit using a SAMPLE (header + first N rows)
    sample_last_row = min(ws.max_row, 1 + max_autofit_rows)  # row 1 is header
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        hv = ws.cell(row=1, column=col_idx).value
        max_len = max(max_len, len("" if hv is None else str(hv)))
        for row in range(2, sample_last_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            max_len = max(max_len, len("" if v is None else str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.read()


def parse_date_yyyy_mm_dd(s: str) -> date:
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def validate_api_window(d_start: date, d_end: date):
    if d_end < d_start:
        return False, "La fecha final no puede ser menor a la fecha inicial."
    days = (d_end - d_start).days + 1
    if days > 31:
        return False, "El rango de fechas no puede ser mayor a 1 mes (máximo 31 días)."
    oldest_allowed = date.today() - timedelta(days=92)
    if d_start < oldest_allowed:
        return False, "La fecha inicial no puede ser anterior a ~3 meses desde hoy (restricción del API)."
    return True, None


def _get_section(*names: str) -> str:
    for n in names:
        if n in st.secrets:
            return n
    raise RuntimeError(f"No encontré ninguna sección en secrets.toml: {names}")


def load_bonsaif_section(section_name: str):
    s = st.secrets[section_name]
    cfg = {
        "BASE_URL": str(s.get("BASE_URL", "https://eva.bonsaif.com/api")).strip(),
        "SERVICE": str(s.get("SERVICE", "cc/api")).strip(),
        "METHOD": str(s.get("METHOD", "27")).strip(),
        "KEY": str(s.get("KEY", "")).strip(),
        "SYS": str(s.get("SYS", "")).strip(),
        "AUTO_FETCH": bool(s.get("AUTO_FETCH", True)),
    }
    if not cfg["KEY"] or not cfg["SYS"]:
        raise RuntimeError(f"{section_name}: KEY/SYS faltantes en secrets.toml")

    campaigns_raw = s.get("CAMPAIGNS", None)
    if not campaigns_raw or not isinstance(campaigns_raw, (list, tuple)):
        raise RuntimeError(f"{section_name}: falta CAMPAIGNS (lista)")

    campaigns = []
    for item in campaigns_raw:
        camp = str(item.get("campana", "")).strip()
        cid = str(item.get("id", "")).strip()
        if not camp or not cid:
            continue
        if not cid.isdigit():
            raise RuntimeError(f"{section_name}: id inválido para '{camp}' (debe ser numérico)")
        campaigns.append({"campana": camp, "id": cid})

    if not campaigns:
        raise RuntimeError(f"{section_name}: no hay campañas válidas en CAMPAIGNS")

    # ✅ DEFAULT DATE RANGE: first day of current month -> today (still clamped to oldest_allowed)
    today = date.today()
    oldest_allowed = today - timedelta(days=92)
    month_start = date(today.year, today.month, 1)

    default_start = max(month_start, oldest_allowed)
    default_end = today

    return cfg, default_start, default_end, campaigns


@st.cache_data(show_spinner=False, ttl=300)
def fetch_campaign(
    base_url: str,
    service: str,
    method: str,
    key: str,
    sys: str,
    fechaini: str,
    fechafin: str,
    campana: str,
    campana_id: str,
):
    params = {
        "service": service,
        "m": method,
        "key": key,
        "sys": sys,
        "fechaini": fechaini,
        "fechafin": fechafin,
        "campana": campana,
        "id": campana_id,
    }
    url = f"{base_url}?{urlencode(params)}"
    r = requests.get(url, timeout=60)
    r.raise_for_status()

    try:
        payload = r.json()
    except Exception:
        return pd.DataFrame(), f"Respuesta no-JSON:\n{r.text[:1500]}"

    msg = _payload_message(payload)
    records = _pick_records_from_payload(payload)
    df = pd.DataFrame.from_records(records) if records else pd.DataFrame()
    df = normalize_df(df)
    return df, msg


def compute_hangup_flag(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip().str.lower()
    return s.isin(["si", "s", "1", "true", "yes", "y"])


def apply_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    out = df.copy()
    for colname, selected in filters.items():
        if colname not in out.columns:
            continue
        if selected:
            out = out[out[colname].astype(str).isin([str(x) for x in selected])]
    return out


def make_pct_table(grouped: pd.DataFrame, group_col: str, cat_col: str, value_col: str = "count") -> pd.DataFrame:
    pivot = grouped.pivot_table(index=group_col, columns=cat_col, values=value_col, aggfunc="sum", fill_value=0)

    # ✅ Rename blank/NaN category headers to "Sin Tipificacion" (and merge if duplicates)
    new_cols = []
    for c in pivot.columns:
        s = "" if c is None else str(c).strip()
        if s == "" or s.lower() == "nan":
            new_cols.append("Sin Tipificacion")
        else:
            new_cols.append(s)
    pivot.columns = new_cols
    if len(set(new_cols)) != len(new_cols):
        pivot = pivot.groupby(level=0, axis=1).sum()

    row_sum = pivot.sum(axis=1).replace(0, 1)
    pct = (pivot.div(row_sum, axis=0) * 100).round(2)
    pct.columns = [f"{c} (%)" for c in pivot.columns]
    return pd.concat([pivot, pct], axis=1).reset_index()


# ----------------------------------------------------
# ✅ ADEUDO parsing from Obs_CC (threshold: < 1500 ADEUDO_TRATABLE)
# ----------------------------------------------------
_RE_MONEY = re.compile(
    r"(?i)(?:\$|mxn|pesos|adeudo|deuda)\s*[:\-]?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)(?:\.[0-9]+)?"
)


def _parse_amount_from_obs(text: str) -> float | None:
    if text is None:
        return None
    t = str(text).strip()
    if not t:
        return None

    m = _RE_MONEY.search(t)
    if m:
        raw = m.group(1).replace(",", "")
        try:
            return float(raw)
        except Exception:
            pass

    nums = re.findall(r"([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)(?:\.[0-9]+)?", t)
    vals = []
    for n in nums:
        try:
            v = float(n.replace(",", ""))
        except Exception:
            continue
        if 1900 <= v <= 2100:
            continue
        if 0 <= v <= 200000:
            vals.append(v)
    if not vals:
        return None
    return max(vals)


def add_adeudo_tratable(
    df: pd.DataFrame,
    col_result: str | None,
    col_obs: str | None,
    threshold: float = 1500.0,
) -> tuple[pd.DataFrame, str | None]:
    if df is None or df.empty or not col_result:
        return df, None

    out = df.copy()
    adj_col = "Codigo_Resultado_Ajustado"
    out[adj_col] = out[col_result].astype(str)

    if not col_obs or col_obs not in out.columns:
        return out, adj_col

    result_norm = out[col_result].astype(str).map(lambda x: _strip_accents(str(x)).lower())
    mask_adeudo = result_norm.str.contains("adeudo", na=False)

    if mask_adeudo.any():
        obs_vals = out.loc[mask_adeudo, col_obs].astype(str)
        amounts = obs_vals.map(_parse_amount_from_obs)

        treatable_mask = amounts.notna() & (amounts < threshold)
        idx_treatable = out.loc[mask_adeudo].index[treatable_mask.values]
        out.loc[idx_treatable, adj_col] = "ADEUDO_TRATABLE"

    return out, adj_col


# ----------------------------------------------------
# ✅ ADD SUPERVISOR COLUMN TO SUMMARY TABLE (Tipificación)
# - NO "Supervisores_n"
# - DO show "SIN_SUPERVISOR" (but never show NaN)
# ----------------------------------------------------
def attach_supervisor_to_tipificacion_table(tbl: pd.DataFrame, df_src: pd.DataFrame, group_col: str) -> pd.DataFrame:
    if tbl is None or tbl.empty:
        return tbl
    if df_src is None or df_src.empty:
        return tbl
    if "Supervisor" not in df_src.columns:
        return tbl
    if group_col not in df_src.columns:
        return tbl
    if group_col == "Supervisor":
        return tbl

    out = tbl.copy()
    src = df_src.copy()

    src["Supervisor"] = _clean_text_to_na(src["Supervisor"]).fillna("SIN_SUPERVISOR").astype(str)

    if group_col == "Campana":
        sup_agg = (
            src.groupby("Campana")["Supervisor"]
            .agg(Supervisor_top=lambda x: (x.mode().iloc[0] if not x.mode().empty else "SIN_SUPERVISOR"))
            .reset_index()
        )
        out = out.merge(sup_agg, on="Campana", how="left")

        front = ["Campana", "Supervisor_top"]
        rest = [c for c in out.columns if c not in front]
        return out[front + rest]

    sup_map = (
        src.groupby(group_col)["Supervisor"]
        .agg(lambda x: (x.mode().iloc[0] if not x.mode().empty else "SIN_SUPERVISOR"))
        .reset_index()
    )
    out = out.merge(sup_map, on=group_col, how="left")

    front = [group_col, "Supervisor"]
    rest = [c for c in out.columns if c not in front]
    return out[front + rest]


# ----------------------------------------------------
# UI
# ----------------------------------------------------
st.title("📞 Bonsaif Tipificación — GLOS")

# ✅ GLOS only (with fallback to BONSAIF if you use that section name)
SOURCE_OPTIONS = {"GLOS": ("BONSAIF_GLOS", "BONSAIF")}

colA, _ = st.columns([1, 2])
with colA:
    source_label = st.selectbox("Fuente", list(SOURCE_OPTIONS.keys()), index=0)

section_name = _get_section(*SOURCE_OPTIONS[source_label])

try:
    cfg, default_start, default_end, campaigns = load_bonsaif_section(section_name)
except Exception as e:
    st.error(str(e))
    st.stop()

today = date.today()
oldest_allowed = today - timedelta(days=92)

# Session state
if "df_all" not in st.session_state:
    st.session_state.df_all = pd.DataFrame()
if "last_ts" not in st.session_state:
    st.session_state.last_ts = None
if "last_query" not in st.session_state:
    st.session_state.last_query = None
if "last_msg" not in st.session_state:
    st.session_state.last_msg = None

# Excel session state (on-demand)
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "excel_key" not in st.session_state:
    st.session_state.excel_key = None

with st.sidebar:
    st.header("Filtros")

    dr = st.date_input(
        "Rango de fechas (máx 31 días, ~3 meses atrás)",
        value=(default_start, default_end),
        min_value=oldest_allowed,
        max_value=today,
    )
    if isinstance(dr, tuple) and len(dr) == 2:
        d_start, d_end = dr
    else:
        d_start = dr
        d_end = dr

    st.markdown("---")
    st.header("Acciones")
    run_btn = st.button("🔄 Actualizar", use_container_width=True)
    clear_btn = st.button("🧹 Limpiar", use_container_width=True)

if clear_btn:
    st.session_state.df_all = pd.DataFrame()
    st.session_state.last_ts = None
    st.session_state.last_query = None
    st.session_state.last_msg = None
    st.session_state.excel_bytes = None
    st.session_state.excel_key = None
    st.rerun()

current_query = (source_label, d_start, d_end)
dates_changed = st.session_state.last_query != current_query
should_fetch = run_btn or (cfg["AUTO_FETCH"] and (st.session_state.last_ts is None or dates_changed))

if should_fetch:
    ok, err = validate_api_window(d_start, d_end)
    if not ok:
        st.error(err)
        st.stop()

    all_dfs = []
    msgs = []
    with st.spinner("Consultando GLOS (todas las campañas)..."):
        for c in campaigns:
            camp = c["campana"]
            cid = c["id"]
            try:
                df_i, msg_i = fetch_campaign(
                    base_url=cfg["BASE_URL"],
                    service=cfg["SERVICE"],
                    method=cfg["METHOD"],
                    key=cfg["KEY"],
                    sys=cfg["SYS"],
                    fechaini=str(d_start),
                    fechafin=str(d_end),
                    campana=camp,
                    campana_id=cid,
                )
                if not df_i.empty:
                    df_i = df_i.copy()
                    df_i["Campana"] = camp
                    df_i["Campana_ID"] = cid
                    all_dfs.append(df_i)
                if msg_i:
                    msgs.append(f"{camp} (ID {cid}): {msg_i}")
            except Exception as e:
                msgs.append(f"{camp} (ID {cid}): ERROR -> {e}")

    df_all = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()
    st.session_state.df_all = df_all
    st.session_state.last_msg = "\n".join(msgs) if msgs else None
    st.session_state.last_ts = datetime.now()
    st.session_state.last_query = current_query

df_all = st.session_state.df_all

if st.session_state.last_ts:
    st.caption(f"Última actualización: {st.session_state.last_ts.strftime('%Y-%m-%d %H:%M:%S')}")

if df_all.empty:
    st.info(st.session_state.last_msg or "No hay datos para mostrar con el rango seleccionado.")
    st.stop()

# Key columns
col_estatus = first_existing_col(df_all, ["Estatus_CC"])
col_result_raw = first_existing_col(df_all, ["Codigo_Resultado_CC"])
col_agent = first_existing_col(df_all, ["Gestor_CC"])
col_hang = first_existing_col(df_all, ["Colgo_Agente_CC"])
col_obs = first_existing_col(df_all, ["Obs_CC", "OBS_CC", "Observaciones", "Observacion"])

# ✅ Supervisor rule for GLOS: Calificacion_Int_CC
col_supervisor_auto = first_existing_col(df_all, ["Calificacion_Int_CC", "Calificacion Int CC", "calificacion_int_cc"])

df = df_all.copy()

# Hangup flag
if col_hang:
    df["Hangup_Flag"] = compute_hangup_flag(df[col_hang])
else:
    df["Hangup_Flag"] = False

# ✅ Supervisor unified (SHOW "SIN_SUPERVISOR", DO NOT SHOW "nan")
if "Supervisor" not in df.columns:
    if col_supervisor_auto and col_supervisor_auto in df.columns:
        df["Supervisor"] = _clean_text_to_na(df[col_supervisor_auto]).fillna("SIN_SUPERVISOR")
    else:
        df["Supervisor"] = "SIN_SUPERVISOR"
else:
    df["Supervisor"] = _clean_text_to_na(df["Supervisor"]).fillna("SIN_SUPERVISOR")

# Adjusted Código Resultado (adeudo rule: <1500 -> ADEUDO_TRATABLE)
df, col_result = add_adeudo_tratable(df, col_result_raw, col_obs, threshold=1500.0)
if not col_result:
    col_result = col_result_raw

# ----------------------------------------------------
# FILTERS
# ----------------------------------------------------
with st.sidebar:
    st.markdown("---")
    st.subheader("Filtros de análisis")

    camp_opts = sorted(df["Campana"].dropna().astype(str).unique().tolist()) if "Campana" in df.columns else []
    selected_campaigns = st.multiselect("Campañas (una o varias)", camp_opts, default=camp_opts)

    sup_opts = sorted(df["Supervisor"].dropna().astype(str).unique().tolist())
    selected_sup = st.multiselect("Supervisor (uno o varios)", sup_opts, default=sup_opts)

    agent_opts = sorted(df[col_agent].dropna().astype(str).unique().tolist()) if col_agent else []
    selected_agents = st.multiselect("Agente (uno o varios)", agent_opts, default=agent_opts)

    estatus_opts = sorted(df[col_estatus].dropna().astype(str).unique().tolist()) if col_estatus else []
    selected_estatus = st.multiselect("Tipificacion (uno o varios)", estatus_opts, default=estatus_opts)

    # Dependent options for Código Resultado
    pre_filters = {}
    if "Campana" in df.columns:
        pre_filters["Campana"] = selected_campaigns
    pre_filters["Supervisor"] = selected_sup
    if col_agent:
        pre_filters[col_agent] = selected_agents
    if col_estatus:
        pre_filters[col_estatus] = selected_estatus

    df_for_results = apply_filters(df, pre_filters)
    result_opts = sorted(df_for_results[col_result].dropna().astype(str).unique().tolist()) if col_result else []
    selected_results = st.multiselect(
        "Subtificación (Código Resultado) — depende de la Tipificacion",
        result_opts,
        default=result_opts,
    )

    hang_filter = st.selectbox("Agente colgó", ["Todos", "Sí", "No"], index=0)

# Apply final filters
filters = {}
if "Campana" in df.columns:
    filters["Campana"] = selected_campaigns
filters["Supervisor"] = selected_sup
if col_agent:
    filters[col_agent] = selected_agents
if col_estatus:
    filters[col_estatus] = selected_estatus
if col_result:
    filters[col_result] = selected_results

df_f = apply_filters(df, filters)

if hang_filter == "Sí":
    df_f = df_f[df_f["Hangup_Flag"] == True]
elif hang_filter == "No":
    df_f = df_f[df_f["Hangup_Flag"] == False]

if df_f.empty:
    st.warning("Con los filtros seleccionados no hay registros.")
    st.stop()

# ----------------------------------------------------
# VIEW MODE
# ----------------------------------------------------
view_mode = st.radio("Vista por", ["Campaña", "Supervisor", "Agente"], horizontal=True)
group_col = {"Campaña": "Campana", "Supervisor": "Supervisor", "Agente": col_agent or "Supervisor"}[view_mode]

# ----------------------------------------------------
# KPIs
# ----------------------------------------------------
st.subheader("📊 KPIs (con filtros actuales)")
k1, k2, k3, k4 = st.columns(4)

k1.metric("Llamadas", f"{len(df_f):,}")
k2.metric("Tipificaciones únicas", f"{df_f[col_estatus].nunique():,}" if col_estatus else "-")
k3.metric("Subtificaciones únicas", f"{df_f[col_result].nunique():,}" if col_result else "-")
hang_rate = (df_f["Hangup_Flag"].mean() * 100) if len(df_f) else 0
k4.metric("Agente colgó (%)", f"{hang_rate:,.2f}%")

# ----------------------------------------------------
# EXCEL KEY (invalidate prepared excel when filters change)
# ----------------------------------------------------
excel_key = (
    source_label,
    str(d_start),
    str(d_end),
    tuple(selected_campaigns),
    tuple(selected_sup),
    tuple(selected_agents),
    tuple(selected_estatus),
    tuple(selected_results),
    hang_filter,
)
if st.session_state.excel_key != excel_key:
    st.session_state.excel_key = excel_key
    st.session_state.excel_bytes = None

# ----------------------------------------------------
# TABS
# ----------------------------------------------------
tab1, tab2, tab3 = st.tabs(
    ["Tipificacion", "Agente colgó", "Detalle + Excel"]
)

with tab1:
    st.markdown("### Análisis por **Tipificacion** (ajustada a filtros)")
    if not col_estatus:
        st.info("No existe la columna Estatus_CC (Tipificacion) en estos datos.")
    else:
        g = df_f.groupby([group_col, col_estatus], as_index=False).size().rename(columns={"size": "count"})
        fig = px.bar(
            g,
            x=group_col,
            y="count",
            color=col_estatus,
            barmode="stack",
            title=f"Tipificacion por {view_mode}",
            labels={col_estatus: "Tipificacion"},
        )
        st.plotly_chart(fig, use_container_width=True)

        tbl = make_pct_table(g, group_col, col_estatus, "count")
        tbl = attach_supervisor_to_tipificacion_table(tbl, df_f, group_col)
        st.dataframe(tbl, use_container_width=True, height=420)

        # ✅ Subtificación BELOW Tipificación (same filters)
        st.markdown("### Subtificación correspondiente (ajustada a filtros)")
        if not col_result:
            st.info("No existe la columna Codigo_Resultado_CC (Subtificación) en estos datos.")
        else:
            g2 = df_f.groupby([group_col, col_result], as_index=False).size().rename(columns={"size": "count"})
            fig2 = px.bar(
                g2,
                x=group_col,
                y="count",
                color=col_result,
                barmode="stack",
                title=f"Subtificación por {view_mode}",
                labels={col_result: "Subtificación"},
            )
            st.plotly_chart(fig2, use_container_width=True)
            st.dataframe(make_pct_table(g2, group_col, col_result, "count"), use_container_width=True, height=420)

with tab2:
    st.markdown("### **Agente colgó** (conteo y %), ajustado a filtros")
    g = df_f.groupby(group_col, as_index=False).agg(total=("Hangup_Flag", "size"), colgo=("Hangup_Flag", "sum"))
    g["pct_colgo"] = (g["colgo"] / g["total"].replace(0, 1) * 100).round(2)
    fig = px.bar(g, x=group_col, y="pct_colgo", title=f"% Agente colgó por {view_mode}")
    st.plotly_chart(fig, use_container_width=True)
    st.dataframe(g.sort_values("pct_colgo", ascending=False), use_container_width=True, height=420)

    # ✅ Subtificación of ONLY hung-up calls
    st.markdown("### Subtificación de **llamadas colgadas** (ajustada a filtros)")
    if not col_result:
        st.info("No existe la columna Codigo_Resultado_CC (Subtificación) en estos datos.")
    else:
        df_h = df_f[df_f["Hangup_Flag"] == True]
        if df_h.empty:
            st.info("No hay llamadas colgadas con los filtros actuales.")
        else:
            g_h = df_h.groupby([group_col, col_result], as_index=False).size().rename(columns={"size": "count"})
            fig_h = px.bar(
                g_h,
                x=group_col,
                y="count",
                color=col_result,
                barmode="stack",
                title=f"Subtificación (solo colgadas) por {view_mode}",
                labels={col_result: "Subtificación"},
            )
            st.plotly_chart(fig_h, use_container_width=True)
            st.dataframe(make_pct_table(g_h, group_col, col_result, "count"), use_container_width=True, height=420)

with tab3:
    st.markdown("### Detalle filtrado")
    st.dataframe(df_f, use_container_width=True, height=520)

    # ✅ ON-DEMAND Excel to avoid Streamlit Cloud crashes
    c1, c2 = st.columns([1, 1])
    with c1:
        prepare_excel = st.button("📦 Preparar Excel (on-demand)", use_container_width=True)
    with c2:
        clear_excel = st.button("🧹 Limpiar Excel preparado", use_container_width=True)

    if clear_excel:
        st.session_state.excel_bytes = None
        st.success("Excel preparado eliminado.")
        st.rerun()

    if prepare_excel:
        if len(df_f) > 60000:
            st.warning("Hay muchos registros. Si tarda o falla, reduce el rango/filters antes de exportar.")
        with st.spinner("Generando Excel..."):
            st.session_state.excel_bytes = to_excel_bytes(df_f, sheet_name="detalle_filtrado", max_autofit_rows=200)
        st.success("Excel listo ✅")

    if st.session_state.excel_bytes:
        st.download_button(
            "⬇️ Descargar Excel (filtrado)",
            data=st.session_state.excel_bytes,
            file_name="bonsaif_m27_GLOS_detalle_filtrado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("Para evitar crashes, el Excel se genera **solo** cuando presionas **Preparar Excel**.")

if st.session_state.last_msg:
    with st.expander("Mensajes del API (por campaña)"):
        st.text(st.session_state.last_msg)
